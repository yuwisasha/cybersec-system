from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models import (
    Incident,
    EventLog,
    Event,
    SeverityLevel,
    IncidentDetail,
    IncidentRecommendation,
)


def generate_incident_report(db: Session) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Инциденты"

    headers = [
        "ID",
        "Дата",
        "Статус",
        "Описание",
        "Критичность",
        "Кол-во событий",
        "Кол-во рекомендаций",
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold_font

    rows = db.query(Incident).order_by(Incident.created_at.desc()).all()

    for incident in rows:
        # критичность — по максимальной среди событий
        severity = (
            db.query(SeverityLevel.name)
            .join(EventLog, EventLog.severity_id == SeverityLevel.id)
            .join(Event, Event.id == EventLog.event_id)
            .join(IncidentDetail, IncidentDetail.event_id == Event.id)
            .filter(IncidentDetail.incident_id == incident.id)
            .order_by(SeverityLevel.id.desc())
            .first()
        )

        event_count = (
            db.query(IncidentDetail).filter_by(incident_id=incident.id).count()
        )
        rec_count = (
            db.query(IncidentRecommendation)
            .filter_by(incident_id=incident.id)
            .count()
        )

        row = [
            incident.id,
            incident.created_at.isoformat(),
            incident.status,
            incident.description,
            severity[0] if severity else "-",
            event_count,
            rec_count,
        ]
        ws.append(row)

        if severity and severity[0].lower() == "критический":
            for cell in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
                for c in cell:
                    c.fill = PatternFill(
                        start_color="FF9999", fill_type="solid"
                    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
